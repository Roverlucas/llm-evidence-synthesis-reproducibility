"""Build the tie-breaker's adjudication workbook for the round-2 residual disagreements.

The senior author adjudicates, per the pre-registration (OSF fgn3e) and protocol §10.
Two constraints from that section shape this file:

  - the tie-break is applied *only* to residual disagreements, against written protocol
    criteria, and **without access to the LLM outputs**. No model decision, score or
    identifier appears anywhere in the workbook.
  - each decision records the criterion invoked and a justification, so the gold standard
    is auditable item by item.

Both raters' positions are shown deliberately. This is adjudication, not a third blind
pass: the tie-breaker is resolving a stated disagreement and needs to see what is being
disagreed about.

Usage:
    python scripts/dual_labeling/export_tiebreak_xlsx.py \
        --audit data/dual_labeling/reconciliation/coordinator_audit_round2.csv \
        --template data/dual_labeling/reconciliation/recalibration_labeler2.csv \
        --out ~/Downloads/tiebreak_Yara_2026-08-22.xlsx
"""
from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DECISIONS = "INCLUDE,EXCLUDE,UNCERTAIN"
CRITERIA = "1,2,3,4,5a,5b,5c,6,nenhum"

INSTRUCTIONS = [
    ("Desempate — subconjunto de validação humana", True),
    ("", False),
    ("Profa. Yara de Souza Tadano — tie-breaker designada no pré-registro OSF fgn3e.", False),
    ("", False),
    ("O QUE É ISTO", True),
    ("Duas pesquisadoras rotularam 100 abstracts de forma independente. Em 25 elas divergiram.", False),
    ("Sob o protocolo v1.2, ambas reavaliaram esses 25 às cegas: 14 convergiram.", False),
    ("Os 11 restantes estão nesta planilha e precisam de uma decisão final.", False),
    ("", False),
    ("POR QUE VOCÊ E NÃO O LUCAS", True),
    ("O Lucas é primeiro autor e construiu o pipeline que é avaliado CONTRA este padrão.", False),
    ("Se ele decidisse os itens contestados, o padrão de referência deixaria de ser", False),
    ("independente do objeto avaliado. O pré-registro previu isso desde maio.", False),
    ("", False),
    ("COMO DECIDIR", True),
    ("Para cada linha: leia o abstract, veja as duas posições, decida e justifique.", False),
    ("Preencha 3 colunas: decisao_final, criterio_invocado, justificativa.", False),
    ("Nenhuma saída de modelo de linguagem aparece aqui, por desenho do protocolo (§10).", False),
    ("", False),
    ("A REGRA QUE ESTÁ EM DISPUTA (protocolo v1.2, §2.1 e §4)", True),
    ("5a  valores numéricos presentes             -> critério ATENDIDO", False),
    ("5b  menciona o efeito, mas sem os valores   -> UNCERTAIN", False),
    ("5c  nenhuma estimativa e nenhuma menção     -> EXCLUDE", False),
    ("", False),
    ("Precedência: EXCLUDE > UNCERTAIN > INCLUDE. Falha em 2 ou mais critérios -> EXCLUDE.", False),
    ("Falha apenas no critério 4 (design) -> UNCERTAIN.", False),
    ("", False),
    ("O QUE OS DADOS MOSTRAM, PARA SUA INFORMAÇÃO", True),
    ("8 dos 11 desacordos são exatamente a fronteira 5b/5c: as duas leem o mesmo abstract", False),
    ("e discordam sobre se ele MENCIONA uma estimativa. A emenda v1.2 desambiguou a regra", False),
    ("escrita, mas não o julgamento sobre o texto. Isso será reportado como resultado.", False),
    ("", False),
    ("Onde o coordenador encontrou algo verificável no próprio registro (não julgamento),", False),
    ("isso está na coluna nota_do_coordenador. São observações, não recomendações.", False),
    ("", False),
    ("Devolver para: lucasrover@alunos.utfpr.edu.br", False),
]


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--audit", required=True, type=Path)
    ap.add_argument("--template", required=True, type=Path)
    ap.add_argument("--out", required=True, type=Path)
    args = ap.parse_args()

    audit = pd.read_csv(args.audit)
    meta = pd.read_csv(args.template).set_index("labeling_id")

    forbidden = [c for c in audit.columns if any(
        k in c.lower() for k in ("emr", "model", "llm", "gpt", "claude", "gemini", "llama")
    )]
    if forbidden:
        raise SystemExit(f"refusing to write: LLM-derived columns present: {forbidden}")

    wb = Workbook()
    ws = wb.active
    ws.title = "instrucoes"
    for i, (line, is_header) in enumerate(INSTRUCTIONS, start=1):
        c = ws.cell(i, 1, line)
        c.font = Font(bold=is_header, size=12 if is_header else 11)
        c.alignment = Alignment(wrap_text=False)
    ws.column_dimensions["A"].width = 96

    sh = wb.create_sheet("desempate")
    cols = ["labeling_id", "pmid", "year", "journal", "title", "abstract",
            "pesquisadora_1_decisao", "pesquisadora_1_criterio", "pesquisadora_1_justificativa",
            "pesquisadora_2_decisao", "pesquisadora_2_criterio", "pesquisadora_2_justificativa",
            "natureza_do_desacordo", "nota_do_coordenador",
            "decisao_final", "criterio_invocado", "justificativa"]
    head = PatternFill("solid", fgColor="1F4E79")
    fill_in = PatternFill("solid", fgColor="C6E0B4")
    for j, name in enumerate(cols, start=1):
        c = sh.cell(1, j, name)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill_in if name in ("decisao_final", "criterio_invocado", "justificativa") else head
        if name in ("decisao_final", "criterio_invocado", "justificativa"):
            c.font = Font(bold=True, color="000000")

    for i, row in enumerate(audit.itertuples(), start=2):
        m = meta.loc[row.labeling_id]
        vals = {
            "labeling_id": row.labeling_id, "pmid": m.pmid, "year": m.year,
            "journal": m.journal, "title": m.title, "abstract": m.abstract,
            "pesquisadora_1_decisao": row.labeler1_decision,
            "pesquisadora_1_criterio": row.labeler1_criteria,
            "pesquisadora_1_justificativa": row.labeler1_rationale,
            "pesquisadora_2_decisao": row.labeler2_decision,
            "pesquisadora_2_criterio": row.labeler2_criteria,
            "pesquisadora_2_justificativa": row.labeler2_rationale,
            "natureza_do_desacordo": row.nature_of_disagreement,
            "nota_do_coordenador": row.coordinator_note,
            "decisao_final": "", "criterio_invocado": "", "justificativa": "",
        }
        for j, name in enumerate(cols, start=1):
            c = sh.cell(i, j, vals[name])
            c.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {"labeling_id": 12, "pmid": 11, "year": 7, "journal": 26, "title": 46,
              "abstract": 88, "natureza_do_desacordo": 16, "nota_do_coordenador": 44,
              "decisao_final": 16, "criterio_invocado": 16, "justificativa": 40}
    for j, name in enumerate(cols, start=1):
        sh.column_dimensions[get_column_letter(j)].width = widths.get(name, 22)
    for i in range(2, len(audit) + 2):
        sh.row_dimensions[i].height = 150
    sh.freeze_panes = "A2"

    dv_d = DataValidation(type="list", formula1=f'"{DECISIONS}"', allow_blank=True, showDropDown=False)
    dv_c = DataValidation(type="list", formula1=f'"{CRITERIA}"', allow_blank=True, showDropDown=False)
    sh.add_data_validation(dv_d); sh.add_data_validation(dv_c)
    col_d = get_column_letter(cols.index("decisao_final") + 1)
    col_c = get_column_letter(cols.index("criterio_invocado") + 1)
    dv_d.add(f"{col_d}2:{col_d}{len(audit)+1}")
    dv_c.add(f"{col_c}2:{col_c}{len(audit)+1}")

    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"rows: {len(audit)}   written: {args.out}")
    print("LLM-derived columns: none (protocol §10)")


if __name__ == "__main__":
    main()
