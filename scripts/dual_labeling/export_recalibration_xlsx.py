"""Export the blinded v1.2 recalibration CSVs as per-labeler XLSX workbooks.

The first round shipped a hand-made pair of workbooks and one labeler received the
wrong template, so the role had to be reassigned after the fact. This script makes
the export reproducible and stamps the role onto the file name, the sheet header
and a dedicated identification block, so a mix-up is visible before any labeling
starts.

Reads ONLY the blinded sheets, which carry no decision from either round — the
recalibration stays independent by construction.

Usage:
    python scripts/dual_labeling/export_recalibration_xlsx.py \
        --source data/dual_labeling/reconciliation/ \
        --out ~/Downloads/recalibracao_v12_<date>/ \
        --labeler1-name Isabelle --labeler2-name Luiza
"""
from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DECISIONS = "INCLUDE,EXCLUDE,UNCERTAIN"
CONFIDENCES = "HIGH,MEDIUM,LOW"

# Metadata is read-only context; the four *_v12 columns are what the labeler fills.
COL_WIDTHS = {
    "labeling_id": 12,
    "pmid": 12,
    "doi": 28,
    "year": 7,
    "journal": 28,
    "title": 60,
    "abstract": 110,
    "url": 34,
}
FILL_WIDTHS = {
    "decision_v12": 16,
    "confidence_v12": 16,
    "rationale_v12": 60,
    "criteria_failed_v12": 20,
}

HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
TOFILL_FILL = PatternFill("solid", fgColor="FFF2CC")

INSTRUCTIONS = [
    ("Rodada de recalibração v1.2 — resumo operacional", True),
    ("", False),
    ("Estes são os 25 abstracts em que houve discordância na primeira rodada.", False),
    ("Os 75 concordantes não são reabertos.", False),
    ("", False),
    ("Reavalie cada um sob o protocolo v1.2 e preencha as quatro colunas *_v12", False),
    ("na aba 'recalibracao'. A rodada continua independente: esta planilha não", False),
    ("mostra a decisão da outra labeler nem a sua decisão anterior.", False),
    ("", False),
    ("Critério 5 — três níveis (a mudança que motivou a rodada)", True),
    ("5a  estimativa numérica E IC 95% numérico          -> critério atendido", False),
    ("5b  diz que estimou o efeito, mas sem os valores   -> UNCERTAIN", False),
    ("5c  nenhuma estimativa e nenhuma menção a uma      -> EXCLUDE", False),
    ("", False),
    ("Peso dos critérios", True),
    ("Falha clara em critério estrutural (1 estudo original, 2 PM2.5,", False),
    ("3 hospitalização respiratória, 6 inglês) -> EXCLUDE, mesmo que seja só um.", False),
    ("Falha apenas no critério 4 (design) ou no 5b -> UNCERTAIN.", False),
    ("Falha em 2 ou mais -> EXCLUDE.", False),
    ("Precedência quando duas regras se aplicam: EXCLUDE > UNCERTAIN > INCLUDE.", False),
    ("", False),
    ("Caso não coberto pela tabela: marque UNCERTAIN, descreva o impasse no", False),
    ("rationale e avise o Lucas. Não improvise.", False),
    ("", False),
    ("Manter a decisão anterior é uma resposta tão útil quanto mudá-la.", False),
    ("Esta rodada não gera um novo κ — não há número a bater.", False),
    ("", False),
    ("O detalhamento completo está em protocolo_v1.2.md (§0, §2.1, §4 e", False),
    ("Exemplos 3, 5 e 6 em §7) e em LEIA-ME_instrucoes.md.", False),
]


def write_workbook(df: pd.DataFrame, prefix: str, name: str, path: Path) -> None:
    role = f"{prefix} = {name}"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="recalibracao", index=False)
        book = writer.book
        sheet = writer.sheets["recalibracao"]

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        for idx, column in enumerate(df.columns, start=1):
            letter = get_column_letter(idx)
            suffix = column.replace(f"{prefix}_", "")
            to_fill = suffix in FILL_WIDTHS
            width = FILL_WIDTHS[suffix] if to_fill else COL_WIDTHS.get(column, 18)
            sheet.column_dimensions[letter].width = width

            header = sheet.cell(row=1, column=idx)
            header.font = Font(bold=True)
            header.fill = TOFILL_FILL if to_fill else HEADER_FILL
            header.alignment = Alignment(vertical="center", wrap_text=True)

            for row in range(2, len(df) + 2):
                cell = sheet.cell(row=row, column=idx)
                cell.alignment = Alignment(vertical="top", wrap_text=column == "abstract")
                if to_fill:
                    cell.fill = TOFILL_FILL

        for suffix, allowed in (("decision_v12", DECISIONS), ("confidence_v12", CONFIDENCES)):
            column = f"{prefix}_{suffix}"
            letter = get_column_letter(df.columns.get_loc(column) + 1)
            validation = DataValidation(
                type="list", formula1=f'"{allowed}"', allow_blank=True, showDropDown=False
            )
            validation.error = f"Valores aceitos: {allowed.replace(',', ', ')}"
            validation.errorTitle = "Fora do vocabulário do protocolo"
            sheet.add_data_validation(validation)
            validation.add(f"{letter}2:{letter}{len(df) + 1}")

        guide = book.create_sheet("instrucoes", 0)
        guide.column_dimensions["A"].width = 96
        guide.cell(row=1, column=1, value=f"Planilha de: {name}  ({prefix})").font = Font(
            bold=True, size=13
        )
        guide.cell(row=2, column=1, value=role).font = Font(italic=True)
        for offset, (text, bold) in enumerate(INSTRUCTIONS, start=4):
            cell = guide.cell(row=offset, column=1, value=text)
            if bold:
                cell.font = Font(bold=True)
        book.active = book.index(book["recalibracao"])


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", required=True, type=Path)
    ap.add_argument("--out", required=True, type=Path)
    ap.add_argument("--labeler1-name", required=True)
    ap.add_argument("--labeler2-name", required=True)
    args = ap.parse_args()

    out = args.out.expanduser()
    out.mkdir(parents=True, exist_ok=True)

    for prefix, name in (("labeler1", args.labeler1_name), ("labeler2", args.labeler2_name)):
        csv_path = args.source / f"recalibration_{prefix}.csv"
        df = pd.read_csv(csv_path, keep_default_na=False)

        own = [c for c in df.columns if c.startswith("labeler")]
        foreign = [c for c in own if not c.startswith(prefix)]
        if foreign:
            raise SystemExit(f"{csv_path}: blinding broken, foreign columns {foreign}")
        filled = [c for c in own if df[c].astype(str).str.strip().ne("").any()]
        if filled:
            raise SystemExit(f"{csv_path}: decision columns are not empty {filled}")

        xlsx_path = out / f"recalibracao_{prefix}_{name}.xlsx"
        write_workbook(df, prefix, name, xlsx_path)
        df.to_csv(out / f"recalibracao_{prefix}_{name}.csv", index=False)
        print(f"{name:10s} {prefix}  {len(df)} linhas  ->  {xlsx_path}")


if __name__ == "__main__":
    main()
